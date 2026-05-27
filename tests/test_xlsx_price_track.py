"""Tests for core/xlsx_price_track.py and xlsx n+ tracking updates."""

from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl

from core.api_client import SymbolHistory
from core.history_cache import HistoryCacheStore, save_symbol_bars, symbol_cache_path
from core.result_xlsx import append_result_xlsx_rows, update_result_price_track
from core.universe_xlsx import append_universe_xlsx_rows, update_universe_price_track
from core.xlsx_price_track import (
    entry_plus_trading_days,
    pct_vs_ref,
    ymd_to_date,
)


def test_pct_vs_ref_rounds_one_decimal():
    assert pct_vs_ref(1050, 1000) == 5.0
    assert pct_vs_ref(950, 1000) == -5.0


def test_entry_plus_trading_days_skips_weekend():
    holidays = frozenset()
    assert entry_plus_trading_days("20260522", 1, holidays) == "20260525"


def test_ymd_to_date_from_date_object():
    assert ymd_to_date(date(2026, 5, 27)) == "20260527"


def _write_history(tmp_path: Path, symbol: str, bars: list[dict]) -> HistoryCacheStore:
    path = symbol_cache_path(tmp_path, symbol)
    save_symbol_bars(path, symbol, bars, full_pages=30)
    return HistoryCacheStore(tmp_path)


def test_universe_price_track_updates_n_plus(tmp_path: Path):
    holiday_path = tmp_path / "holidays.txt"
    holiday_path.write_text("", encoding="utf-8")
    xlsx_path = tmp_path / "universe.xlsx"
    append_universe_xlsx_rows(
        xlsx_path,
        [
            {
                "date_kst": "20260526",
                "strategy_mode": 1,
                "symbol": "005930",
                "name": "Samsung",
                "market_cap_billion": 500000,
                "rs_return_pct": 25.0,
                "rs_rank": 1,
                "w52_high": 1000,
                "vol_ma20": 100,
                "close": 990,
            }
        ],
    )

    store = _write_history(
        tmp_path / "hc",
        "005930",
        [
            {"date": "20260527", "open": 1, "high": 1100, "low": 1, "close": 2, "volume": 1},
        ],
    )
    updated = update_universe_price_track(xlsx_path, "20260527", store, holiday_path)
    assert updated == 1

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    assert ws.cell(3, 12).value == 1000
    assert ws.cell(3, 13).value == 10.0
    assert ws.cell(3, 13).font.color.rgb in ("00C00000", "C00000")


def test_result_price_track_updates_n_plus(tmp_path: Path):
    holiday_path = tmp_path / "holidays.txt"
    holiday_path.write_text("", encoding="utf-8")
    xlsx_path = tmp_path / "result.xlsx"
    from datetime import datetime
    from zoneinfo import ZoneInfo

    buy_ts = datetime(2026, 5, 26, 9, 44, 32, tzinfo=ZoneInfo("Asia/Seoul"))
    append_result_xlsx_rows(
        xlsx_path,
        [
            {
                "kind": "OPEN",
                "symbol": "380540",
                "buy_ts_last": buy_ts,
                "buy_avg": 5550.0,
                "buy_qty": 1,
                "buy_amt": 5550.0,
                "fee": 0.0,
                "tax": 0.0,
            }
        ],
        {"380540": "Opticore"},
    )

    store = _write_history(
        tmp_path / "hc",
        "380540",
        [
            {"date": "20260527", "open": 1, "high": 6000, "low": 1, "close": 2, "volume": 1},
        ],
    )
    updated = update_result_price_track(xlsx_path, "20260527", store, holiday_path)
    assert updated == 1

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    assert ws.cell(3, 22).value == 5550
    assert ws.cell(3, 23).value == 8.1
