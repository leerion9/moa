"""Tests for gap_result.xlsx writer."""

from __future__ import annotations

from pathlib import Path

import openpyxl

from core.gap_collector_logic import hhmmss_to_time
from core.gap_result_xlsx import (
    HEADERS,
    append_gap_result_rows,
    read_existing_buy_dates,
    read_last_cumulative_pnl,
    read_max_no,
    sheet_name_for_year,
)


def test_sheet_name_for_year():
    assert sheet_name_for_year("20260609") == "2026"


def test_append_gap_result_rows(tmp_path: Path):
    path = tmp_path / "gap_result.xlsx"
    rows = [
        {
            "buy_ymd": "20260609",
            "sell_ymd": "20260609",
            "buy_time": hhmmss_to_time("102100"),
            "sell_time": hhmmss_to_time("143000"),
            "symbol": "417010",
            "name": "나노팀",
            "gap_pct": 3.5,
            "max_dip_pct": 5.0,
            "buy_price": 15250,
            "sell_price": 16000,
            "qty": 1,
            "buy_amount": 15250.0,
            "sell_amount": 16000.0,
            "pnl": 750.0,
            "return_pct": 4.91,
            "tax": 28.8,
            "fee_total": 4.8,
            "market_cap_billion": 3651,
            "trading_value_won": 30_000_000_000,
            "trading_value_billion": 300,
            "daily_volume": 1_234_567,
            "approx_trading_value_billion": 187,
            "shares_outstanding": 239_000_000,
            "approx_market_cap_billion": 3648,
        }
    ]
    n = append_gap_result_rows(path, rows, include_market_fields=True)
    assert n == 1
    assert "20260609" in read_existing_buy_dates(path)
    assert read_max_no(path) == 1
    assert read_last_cumulative_pnl(path) == 750.0

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["2026"]
    assert ws.cell(2, 1).value == HEADERS[0]
    assert ws.cell(3, 6).value == 417010
    assert ws.cell(3, 21).value == 1
    assert ws.cell(3, 26).value == 1_234_567
    assert ws.cell(3, 27).value == 187
    assert ws.cell(3, 28).value == 239_000_000
    assert ws.cell(3, 29).value == 3648
    wb.close()

    rows2 = [
        {
            **rows[0],
            "symbol": "005930",
            "name": "삼성전자",
            "pnl": -100.0,
            "return_pct": -1.0,
            "market_cap_billion": 5000000,
            "trading_value_won": 1_000_000_000_000,
            "trading_value_billion": 10000,
        }
    ]
    append_gap_result_rows(path, rows2, include_market_fields=True)
    assert read_max_no(path) == 2
    assert read_last_cumulative_pnl(path) == 650.0


def test_append_without_market_fields(tmp_path: Path):
    path = tmp_path / "gap_backfill.xlsx"
    rows = [
        {
            "buy_ymd": "20250609",
            "sell_ymd": "20250609",
            "buy_time": hhmmss_to_time("110000"),
            "sell_time": hhmmss_to_time("150000"),
            "symbol": "005930",
            "name": "삼성전자",
            "gap_pct": 4.0,
            "max_dip_pct": 3.5,
            "buy_price": 70000,
            "sell_price": 71000,
            "qty": 1,
            "buy_amount": 70000.0,
            "sell_amount": 71000.0,
            "pnl": 900.0,
            "return_pct": 1.28,
            "tax": 100.0,
            "fee_total": 20.0,
            "daily_volume": 500_000,
            "approx_trading_value_billion": 350,
            "shares_outstanding": 800_000_000,
            "approx_market_cap_billion": 56000,
        }
    ]
    append_gap_result_rows(path, rows, include_market_fields=False)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["2025"]
    assert ws.cell(3, 23).value is None
    assert ws.cell(3, 24).value is None
    assert ws.cell(3, 26).value == 500_000
    assert ws.cell(3, 27).value == 350
    assert ws.cell(3, 28).value == 800_000_000
    assert ws.cell(3, 29).value == 56000
    wb.close()
