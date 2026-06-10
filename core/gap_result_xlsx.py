"""xlsx writer for gap-up recovery strategy daily/backfill results."""

from __future__ import annotations

from datetime import date, time
from pathlib import Path
from typing import Any, Dict, List, Optional, Set

import openpyxl

from core.gap_collector_logic import ymd_to_date
from core.vi_collector_logic import cap_vs_trading_value_pct, trading_value_to_billion_won
from core.xlsx_price_track import write_pct_cell

HEADERS = [
    "번호",
    "매수날짜",
    "매수시간",
    "매도날짜",
    "매도시간",
    "종목코드",
    "종목명",
    "전일종가대비시가갭",
    "당일시가대비하락",
    "매수단가",
    "매수수량",
    "총매수금액",
    "매도단가",
    "매도수량",
    "총매도금액",
    "손익",
    "수익률",
    "세금",
    "수수료",
    "누적손익",
    "성공여부",
    "평균수익률",
    "종가매도수익률",
    "익일시가매도수익률",
    "종가매도여부",
    "시총",
    "거래대금",
    "시총대비거래대금비",
    "당일거래량",
    "거래대금(근사)",
    "발행주식수(현재)",
    "시총(근사)",
]

_COL = {h: i for i, h in enumerate(HEADERS, 1)}

_ROW1_HINTS: Dict[int, str] = {
    8: "%",
    9: "%",
    17: "%",
    22: "%",
    23: "%",
    24: "%",
    26: "억원",
    27: "억원",
    28: "%",
    29: "주",
    30: "억원",
    31: "주",
    32: "억원",
}


def sheet_name_for_year(ymd: str) -> str:
    text = str(ymd or "").strip().replace("-", "").replace(".", "")
    return text[:4] if len(text) >= 4 else "0000"


def ensure_gap_sheet(ws) -> None:
    for col_idx, hint in _ROW1_HINTS.items():
        ws.cell(1, col_idx).value = hint
    for col_idx, h in enumerate(HEADERS, 1):
        ws.cell(2, col_idx).value = h


def ensure_gap_result_xlsx(path: Path, year: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        wb = openpyxl.load_workbook(path)
        if year not in wb.sheetnames:
            ws = wb.create_sheet(year)
            ensure_gap_sheet(ws)
            wb.save(path)
        else:
            ws = wb[year]
            ensure_gap_sheet(ws)
            wb.save(path)
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = year
    ensure_gap_sheet(ws)
    wb.save(path)


def _find_last_data_row(ws) -> int:
    last = 2
    for row_idx in range(3, ws.max_row + 1):
        if ws.cell(row_idx, _COL["번호"]).value is not None:
            last = row_idx
    return last


def read_max_no(path: Path) -> int:
    if not path.exists():
        return 0
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        max_no = 0
        for sn in wb.sheetnames:
            ws = wb[sn]
            for row_idx in range(3, ws.max_row + 1):
                val = ws.cell(row_idx, _COL["번호"]).value
                if val is None:
                    continue
                try:
                    max_no = max(max_no, int(val))
                except Exception:
                    continue
        return max_no
    except Exception:
        return 0


def read_last_cumulative_pnl(path: Path) -> float:
    if not path.exists():
        return 0.0
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        best_row = None
        best_no = -1
        for sn in wb.sheetnames:
            ws = wb[sn]
            for row_idx in range(3, ws.max_row + 1):
                no_val = ws.cell(row_idx, _COL["번호"]).value
                if no_val is None:
                    continue
                try:
                    no = int(no_val)
                except Exception:
                    continue
                if no > best_no:
                    best_no = no
                    best_row = ws.cell(row_idx, _COL["누적손익"]).value
        if best_row is None:
            return 0.0
        return float(best_row)
    except Exception:
        return 0.0


def read_existing_trade_keys(path: Path) -> Set[tuple[str, str]]:
    """(종목코드, 매수날짜 YYYYMMDD) pairs already in workbook."""
    if not path.exists():
        return set()
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        out: Set[tuple[str, str]] = set()
        sym_col = _COL["종목코드"]
        date_col = _COL["매수날짜"]
        for sn in wb.sheetnames:
            ws = wb[sn]
            for row_idx in range(3, ws.max_row + 1):
                sym_val = ws.cell(row_idx, sym_col).value
                date_val = ws.cell(row_idx, date_col).value
                if sym_val is None or date_val is None:
                    continue
                sym = str(sym_val).strip().zfill(6)
                if hasattr(date_val, "strftime"):
                    ymd = date_val.strftime("%Y%m%d")
                else:
                    ymd = str(date_val).strip().replace("-", "").replace(".", "")
                if len(sym) == 6 and len(ymd) == 8 and ymd.isdigit():
                    out.add((sym, ymd))
        return out
    except Exception:
        return set()


def read_existing_buy_dates(path: Path) -> Set[str]:
    if not path.exists():
        return set()
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        out: Set[str] = set()
        col = _COL["매수날짜"]
        for sn in wb.sheetnames:
            ws = wb[sn]
            for row_idx in range(3, ws.max_row + 1):
                val = ws.cell(row_idx, col).value
                if val is None:
                    continue
                if hasattr(val, "strftime"):
                    out.add(val.strftime("%Y%m%d"))
                    continue
                text = str(val).strip().replace("-", "").replace(".", "")
                if len(text) == 8 and text.isdigit():
                    out.add(text)
        return out
    except Exception:
        return set()


def read_last_avg_return(path: Path) -> tuple[int, float]:
    """Returns (trade_count, running_avg_return_pct) from the last data row."""
    count = read_max_no(path)
    if count <= 0 or not path.exists():
        return 0, 0.0
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        best_no = -1
        avg_ret = 0.0
        for sn in wb.sheetnames:
            sheet = wb[sn]
            for row_idx in range(3, sheet.max_row + 1):
                no_val = sheet.cell(row_idx, _COL["번호"]).value
                if no_val is None:
                    continue
                try:
                    no = int(no_val)
                except Exception:
                    continue
                if no > best_no:
                    best_no = no
                    avg_ret = float(sheet.cell(row_idx, _COL["평균수익률"]).value or 0)
        return count, avg_ret
    except Exception:
        return 0, 0.0


def _running_avg(prev_avg: float, prev_count: int, value: float) -> float:
    if prev_count <= 0:
        return round(value, 2)
    total = prev_avg * prev_count + value
    return round(total / (prev_count + 1), 2)


def append_gap_result_rows(
    path: Path,
    rows: List[Dict[str, Any]],
    *,
    include_market_fields: bool = True,
) -> int:
    if not rows:
        return 0

    year = sheet_name_for_year(str(rows[0].get("buy_ymd", "") or ""))
    ensure_gap_result_xlsx(path, year)

    max_no = read_max_no(path)
    cum_pnl = read_last_cumulative_pnl(path)
    prev_count, prev_avg_ret = read_last_avg_return(path)

    wb = openpyxl.load_workbook(path)
    ws = wb[year]
    ensure_gap_sheet(ws)
    start_row = _find_last_data_row(ws) + 1

    for i, r in enumerate(rows):
        row_num = start_row + i
        no = max_no + i + 1

        buy_ymd = str(r.get("buy_ymd", "") or "")
        buy_date = ymd_to_date(buy_ymd) if buy_ymd else None
        sell_ymd = str(r.get("sell_ymd", buy_ymd) or buy_ymd)
        sell_date = ymd_to_date(sell_ymd) if sell_ymd else buy_date

        pnl = float(r.get("pnl", 0) or 0)
        cum_pnl = round(cum_pnl + pnl, 1)
        ret_pct = r.get("return_pct")
        success = 1 if pnl > 0 else 0

        sym = str(r.get("symbol", "") or "").strip()
        sym_cell: Any = int(sym) if sym.isdigit() else sym

        ws.cell(row_num, _COL["번호"]).value = no
        ws.cell(row_num, _COL["매수날짜"]).value = buy_date
        ws.cell(row_num, _COL["매수시간"]).value = r.get("buy_time")
        ws.cell(row_num, _COL["매도날짜"]).value = sell_date
        ws.cell(row_num, _COL["매도시간"]).value = r.get("sell_time")
        ws.cell(row_num, _COL["종목코드"]).value = sym_cell
        ws.cell(row_num, _COL["종목명"]).value = r.get("name") or ""
        write_pct_cell(ws.cell(row_num, _COL["전일종가대비시가갭"]), r.get("gap_pct"))
        write_pct_cell(ws.cell(row_num, _COL["당일시가대비하락"]), r.get("max_dip_pct"))
        ws.cell(row_num, _COL["매수단가"]).value = r.get("buy_price")
        ws.cell(row_num, _COL["매수수량"]).value = r.get("qty")
        ws.cell(row_num, _COL["총매수금액"]).value = r.get("buy_amount")
        ws.cell(row_num, _COL["매도단가"]).value = r.get("sell_price")
        ws.cell(row_num, _COL["매도수량"]).value = r.get("qty")
        ws.cell(row_num, _COL["총매도금액"]).value = r.get("sell_amount")
        ws.cell(row_num, _COL["손익"]).value = pnl
        write_pct_cell(ws.cell(row_num, _COL["수익률"]), ret_pct)
        ws.cell(row_num, _COL["세금"]).value = r.get("tax")
        ws.cell(row_num, _COL["수수료"]).value = r.get("fee_total")
        ws.cell(row_num, _COL["누적손익"]).value = cum_pnl
        ws.cell(row_num, _COL["성공여부"]).value = success

        ret_val = float(ret_pct or 0)
        prev_count += 1
        avg_ret = _running_avg(prev_avg_ret, prev_count - 1, ret_val)
        prev_avg_ret = avg_ret
        write_pct_cell(ws.cell(row_num, _COL["평균수익률"]), avg_ret)

        close_only = float(r.get("close_only_return_pct", ret_val) or ret_val)
        write_pct_cell(ws.cell(row_num, _COL["종가매도수익률"]), close_only)

        next_open_pct = r.get("next_open_return_pct")
        if next_open_pct is not None:
            write_pct_cell(
                ws.cell(row_num, _COL["익일시가매도수익률"]),
                float(next_open_pct),
            )

        close_sell = r.get("close_sell")
        if close_sell is None:
            reason = str(r.get("sell_reason", "") or "").strip().lower()
            close_sell = 1 if reason == "close" else 0
        ws.cell(row_num, _COL["종가매도여부"]).value = int(close_sell)

        if include_market_fields:
            cap = r.get("market_cap_billion")
            tv_billion = r.get("trading_value_billion")
            tv_won = int(r.get("trading_value_won", 0) or 0)
            cap_int = int(cap) if cap else None
            ws.cell(row_num, _COL["시총"]).value = cap_int
            ws.cell(row_num, _COL["거래대금"]).value = tv_billion
            write_pct_cell(
                ws.cell(row_num, _COL["시총대비거래대금비"]),
                cap_vs_trading_value_pct(cap_int, tv_won),
            )

        daily_volume = r.get("daily_volume")
        if daily_volume is not None:
            ws.cell(row_num, _COL["당일거래량"]).value = int(daily_volume)
        approx_tv_billion = r.get("approx_trading_value_billion")
        if approx_tv_billion is not None:
            ws.cell(row_num, _COL["거래대금(근사)"]).value = int(approx_tv_billion)
        shares = r.get("shares_outstanding")
        if shares is not None:
            ws.cell(row_num, _COL["발행주식수(현재)"]).value = int(shares)
        approx_cap = r.get("approx_market_cap_billion")
        if approx_cap is not None:
            ws.cell(row_num, _COL["시총(근사)"]).value = int(approx_cap)

    wb.save(path)
    return len(rows)
