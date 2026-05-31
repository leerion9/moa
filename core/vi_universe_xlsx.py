"""xlsx writer for daily static upward VI records."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional, Set

import openpyxl

from core.vi_collector_logic import (
    ViEventRow,
    cap_vs_trading_value_pct,
    format_hhmmss_display,
    trading_value_to_billion_won,
    trigger_vs_release_pct,
)
from core.xlsx_price_track import write_pct_cell

HEADERS = [
    "번호",
    "날짜",
    "vi_발동시간",
    "vi_해제시각",
    "2차상승vi_여부",
    "종목코드",
    "종목명",
    "시장구분",
    "vi_발동가",
    "vi_해제가",
    "발동vs해제",
    "해제후_최고상승_pct",
    "해제후_최저하락_pct",
    "시가총액_억",
    "vi전_거래대금",
    "시총vs거래대금",
    "시총그룹",
]

_SHEET_NAME = "vi_universe_all"
_COL = {h: i for i, h in enumerate(HEADERS, 1)}


def ensure_vi_universe_xlsx(path: Path) -> None:
    if path.exists():
        migrate_vi_universe_xlsx(path)
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _SHEET_NAME
    for col_idx, h in enumerate(HEADERS, 1):
        ws.cell(2, col_idx).value = h
    wb.save(path)


def _find_last_data_row(ws) -> int:
    last = 2
    for row_idx in range(3, ws.max_row + 1):
        if ws.cell(row_idx, 1).value is not None:
            last = row_idx
    return last


def _safe_int(value: object) -> int:
    try:
        if value is None:
            return 0
        text = str(value).replace(",", "").strip()
        if not text:
            return 0
        return int(float(text))
    except Exception:
        return 0


def _write_vi_derived_cells(ws, row_num: int, row: Dict[str, Any]) -> None:
    trigger = _safe_int(row.get("trigger_price"))
    release = _safe_int(row.get("release_price"))
    cap = row.get("market_cap_billion")
    cap_int = int(cap) if cap else None
    tv_raw = row.get("pre_vi_trading_value")
    tv_won = int(tv_raw) if tv_raw else None
    if tv_won is not None and tv_won <= 0:
        tv_won = None

    ws.cell(row_num, _COL["vi_발동가"]).value = trigger if trigger > 0 else None
    ws.cell(row_num, _COL["vi_해제가"]).value = release if release > 0 else None
    write_pct_cell(
        ws.cell(row_num, _COL["발동vs해제"]),
        trigger_vs_release_pct(trigger, release),
    )
    write_pct_cell(
        ws.cell(row_num, _COL["해제후_최고상승_pct"]),
        row.get("post_release_high_pct"),
    )
    write_pct_cell(
        ws.cell(row_num, _COL["해제후_최저하락_pct"]),
        row.get("post_release_low_pct"),
    )
    ws.cell(row_num, _COL["시가총액_억"]).value = cap_int
    tv_billion = trading_value_to_billion_won(tv_won)
    ws.cell(row_num, _COL["vi전_거래대금"]).value = tv_billion
    write_pct_cell(
        ws.cell(row_num, _COL["시총vs거래대금"]),
        cap_vs_trading_value_pct(cap_int, tv_won),
    )
    ws.cell(row_num, _COL["시총그룹"]).value = row.get("cap_group") or ""


def migrate_vi_universe_xlsx(path: Path) -> int:
    """Add/repair derived columns on existing vi_universe.xlsx. Returns updated row count."""
    if not path.exists():
        return 0

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    for col_idx, h in enumerate(HEADERS, 1):
        ws.cell(2, col_idx).value = h

    updated = 0
    for row_idx in range(3, ws.max_row + 1):
        if ws.cell(row_idx, _COL["번호"]).value is None:
            continue

        trigger = _safe_int(ws.cell(row_idx, _COL["vi_발동가"]).value)
        release = _safe_int(ws.cell(row_idx, _COL["vi_해제가"]).value)
        cap_raw = ws.cell(row_idx, _COL["시가총액_억"]).value
        cap_int = _safe_int(cap_raw) if cap_raw is not None else None
        if cap_int is not None and cap_int <= 0:
            cap_int = None

        tv_cell = ws.cell(row_idx, _COL["vi전_거래대금"]).value
        tv_won: Optional[int] = None
        if tv_cell is not None:
            tv_int = _safe_int(tv_cell)
            if tv_int >= 100_000_000:
                tv_won = tv_int
            elif tv_int > 0:
                tv_won = tv_int * 100_000_000

        row = {
            "trigger_price": trigger,
            "release_price": release,
            "post_release_high_pct": ws.cell(row_idx, _COL["해제후_최고상승_pct"]).value,
            "post_release_low_pct": ws.cell(row_idx, _COL["해제후_최저하락_pct"]).value,
            "market_cap_billion": cap_int,
            "pre_vi_trading_value": tv_won,
            "cap_group": ws.cell(row_idx, _COL["시총그룹"]).value,
        }
        _write_vi_derived_cells(ws, row_idx, row)
        updated += 1

    wb.save(path)
    return updated


def read_max_no(path: Path) -> int:
    if not path.exists():
        return 0
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        max_no = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] is None:
                continue
            try:
                max_no = max(max_no, int(row[0]))
            except Exception:
                continue
        return max_no
    except Exception:
        return 0


def read_existing_dates(path: Path) -> Set[str]:
    if not path.exists():
        return set()
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        out: Set[str] = set()
        col = _COL["날짜"]
        for row_idx in range(3, ws.max_row + 1):
            val = ws.cell(row_idx, col).value
            if val is None:
                continue
            text = str(val).strip().replace("-", "").replace(".", "")
            if len(text) == 8 and text.isdigit():
                out.add(text)
        return out
    except Exception:
        return set()


def vi_event_to_row_dict(date_kst: str, event: ViEventRow) -> Dict[str, Any]:
    return {
        "date_kst": date_kst,
        "trigger_time": format_hhmmss_display(event.trigger_hhmmss),
        "release_time": format_hhmmss_display(event.release_hhmmss),
        "has_second_vi": "Y" if event.has_second_vi else "N",
        "symbol": event.symbol,
        "name": event.name,
        "market": event.market,
        "trigger_price": event.trigger_price,
        "release_price": event.release_price,
        "post_release_high_pct": event.post_release_high_pct,
        "post_release_low_pct": event.post_release_low_pct,
        "market_cap_billion": event.market_cap_billion,
        "pre_vi_trading_value": event.pre_vi_trading_value,
        "cap_group": event.cap_group,
    }


def append_vi_universe_xlsx_rows(
    path: Path,
    date_kst: str,
    events: List[ViEventRow],
) -> int:
    if not events:
        return 0
    ensure_vi_universe_xlsx(path)
    max_no = read_max_no(path)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    start_row = _find_last_data_row(ws) + 1

    for i, event in enumerate(events):
        row_num = start_row + i
        no = max_no + i + 1
        r = vi_event_to_row_dict(date_kst, event)
        sym = str(r["symbol"])
        sym_cell: Any = int(sym) if sym.isdigit() else sym

        ws.cell(row_num, _COL["번호"]).value = no
        ws.cell(row_num, _COL["날짜"]).value = date_kst
        ws.cell(row_num, _COL["vi_발동시간"]).value = r["trigger_time"]
        ws.cell(row_num, _COL["vi_해제시각"]).value = r["release_time"]
        ws.cell(row_num, _COL["2차상승vi_여부"]).value = r["has_second_vi"]
        ws.cell(row_num, _COL["종목코드"]).value = sym_cell
        ws.cell(row_num, _COL["종목명"]).value = r["name"]
        ws.cell(row_num, _COL["시장구분"]).value = r["market"]
        _write_vi_derived_cells(ws, row_num, r)

    wb.save(path)
    return len(events)


def remove_vi_universe_rows_by_date(path: Path, date_kst: str) -> int:
    """Remove all data rows for date_kst and renumber. Returns removed count."""
    if not path.exists():
        return 0
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    date_col = _COL["날짜"]
    keep_rows: List[List[Any]] = []
    removed = 0
    for row_idx in range(3, ws.max_row + 1):
        val = ws.cell(row_idx, date_col).value
        text = str(val or "").strip().replace("-", "").replace(".", "")
        if text == date_kst:
            removed += 1
            continue
        if ws.cell(row_idx, _COL["번호"]).value is None:
            continue
        keep_rows.append(
            [ws.cell(row_idx, col).value for col in range(1, len(HEADERS) + 1)]
        )

    if removed <= 0:
        return 0

    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)

    for i, row_vals in enumerate(keep_rows):
        row_num = 3 + i
        for col, value in enumerate(row_vals, 1):
            ws.cell(row_num, col).value = value

    for i in range(len(keep_rows)):
        ws.cell(3 + i, _COL["번호"]).value = i + 1

    wb.save(path)
    return removed
