"""xlsx writer for daily universe watchlist records."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional, TYPE_CHECKING

import openpyxl

from core.trading_day import load_manual_holiday_set
from core.xlsx_price_track import (
    TRACK_HEADERS,
    entry_plus_trading_days,
    get_daily_high_from_bars,
    normalize_symbol,
    pct_vs_ref,
    write_pct_cell,
    write_ref_price_cell,
    ymd_to_date,
)

if TYPE_CHECKING:
    from core.history_cache import HistoryCacheStore

HEADERS = [
    "번호",
    "날짜",
    "전략",
    "종목코드",
    "종목명",
    "시가총액_억",
    "RS수익률_pct",
    "RS순위",
    "52주고가",
    "vol_ma20",
    "전일종가",
] + TRACK_HEADERS

_SHEET_NAME = "universe_all"
_COL = {h: i for i, h in enumerate(HEADERS, 1)}


def ensure_universe_xlsx(path: Path) -> None:
    if path.exists():
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _SHEET_NAME
    for col_idx, h in enumerate(HEADERS, 1):
        ws.cell(2, col_idx).value = h
    wb.save(path)


def _find_last_data_row(ws) -> int:
    last = 2
    for row_idx in range(3, ws.max_row + 1):
        if ws.cell(row_idx, 1).value is not None:
            last = row_idx
    return last


def read_max_no(path: Path) -> int:
    if not path.exists():
        return 0
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        max_no = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] is None:
                continue
            try:
                max_no = max(max_no, int(row[0]))
            except Exception:
                continue
        return max_no
    except Exception:
        return 0


def append_universe_xlsx_rows(
    path: Path,
    rows: List[Dict[str, Any]],
    symbol_names: Optional[Dict[str, str]] = None,
) -> None:
    if not rows:
        return
    ensure_universe_xlsx(path)
    names = symbol_names or {}
    max_no = read_max_no(path)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    start_row = _find_last_data_row(ws) + 1

    for i, r in enumerate(rows):
        row_num = start_row + i
        no = max_no + i + 1
        sym_raw = str(r.get("symbol", "") or "").strip()
        sym = sym_raw.zfill(6) if sym_raw.isdigit() else sym_raw
        name = str(r.get("name", "") or names.get(sym, "") or "").strip()
        sym_cell: Any = int(sym) if sym.isdigit() else sym
        w52_high = int(r.get("w52_high", 0) or 0)

        ws.cell(row_num, _COL["번호"]).value = no
        ws.cell(row_num, _COL["날짜"]).value = str(r.get("date_kst", "") or "")
        ws.cell(row_num, _COL["전략"]).value = int(r.get("strategy_mode", 0) or 0)
        ws.cell(row_num, _COL["종목코드"]).value = sym_cell
        ws.cell(row_num, _COL["종목명"]).value = name
        ws.cell(row_num, _COL["시가총액_억"]).value = r.get("market_cap_billion")
        rs_pct = r.get("rs_return_pct")
        ws.cell(row_num, _COL["RS수익률_pct"]).value = (
            round(float(rs_pct), 4) if rs_pct is not None else None
        )
        ws.cell(row_num, _COL["RS순위"]).value = r.get("rs_rank")
        ws.cell(row_num, _COL["52주고가"]).value = w52_high
        ws.cell(row_num, _COL["vol_ma20"]).value = int(r.get("vol_ma20", 0) or 0)
        ws.cell(row_num, _COL["전일종가"]).value = int(r.get("close", 0) or 0)
        write_ref_price_cell(ws.cell(row_num, _COL["n"]), w52_high)

    wb.save(path)


def update_universe_price_track(
    path: Path,
    as_of_ymd: str,
    history_store: "HistoryCacheStore",
    holiday_path: Path,
) -> int:
    """Fill/backfill n+1..n+15 for watchlist rows. Returns updated cell count."""
    if not path.exists():
        return 0

    holidays = load_manual_holiday_set(holiday_path)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    updated = 0

    for row_idx in range(3, ws.max_row + 1):
        if ws.cell(row_idx, _COL["번호"]).value is None:
            continue

        entry_ymd = ymd_to_date(ws.cell(row_idx, _COL["날짜"]).value)
        symbol = normalize_symbol(ws.cell(row_idx, _COL["종목코드"]).value)
        if not entry_ymd or not symbol:
            continue

        ref_raw = ws.cell(row_idx, _COL["n"]).value
        if ref_raw is None:
            ref_raw = ws.cell(row_idx, _COL["52주고가"]).value
        try:
            ref = int(round(float(ref_raw or 0)))
        except Exception:
            ref = 0
        if ref <= 0:
            continue

        hist = history_store.load_history(symbol)
        bars = list(hist.bars) if hist and hist.bars else []

        for k in range(1, 16):
            target_ymd = entry_plus_trading_days(entry_ymd, k, holidays)
            if target_ymd > as_of_ymd:
                break
            high = get_daily_high_from_bars(bars, target_ymd)
            if high is None:
                continue
            pct = pct_vs_ref(high, ref)
            col = _COL[f"n+{k}"]
            write_pct_cell(ws.cell(row_idx, col), pct)
            updated += 1

    wb.save(path)
    return updated
