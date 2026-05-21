"""
xlsx writer for trade results.

reasult_01.xlsx 양식 기준 구조:
  Row 1  : A-R 비어있음, S=AVERAGE(S3:S9978), T=AVERAGE(T3:T9978)
  Row 2  : 헤더 (번호 ~ 누적손익, 성공률, 평균수익률)
  Row 3+ : 데이터
           R열(누적손익): =R{row-1}+N{row} 공식 (Excel 계산)
           S열(성공률):  =IF(O{row}>=0,1,0)
           T열(평균수익률): =O{row}/100
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional
from zoneinfo import ZoneInfo

import openpyxl

KST = ZoneInfo("Asia/Seoul")

HEADERS = [
    "번호", "매수날짜", "매수시간", "매도날짜", "매도시간",
    "종목코드", "종목명", "매수단가", "매수수량", "총매수금액",
    "매도단가", "매도수량", "총매도금액", "손익", "수익률",
    "세금", "수수료", "누적손익", "성공률", "평균수익률", "매매전략",
]

_SHEET_NAME = "result_all"

# 열 인덱스 (1-based, openpyxl 기준)
_COL = {h: i for i, h in enumerate(HEADERS, 1)}


def ensure_result_xlsx(path: Path) -> None:
    """파일이 없을 때만 신규 생성 (AVERAGE 공식 행 + 헤더 행)."""
    if path.exists():
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _SHEET_NAME
    # Row 1: S/T열 AVERAGE 공식
    ws.cell(1, _COL["성공률"]).value = "=AVERAGE(S3:S9978)"
    ws.cell(1, _COL["평균수익률"]).value = "=AVERAGE(T3:T9978)"
    # Row 2: 헤더
    for col_idx, h in enumerate(HEADERS, 1):
        ws.cell(2, col_idx).value = h
    wb.save(path)


def _find_last_data_row(ws) -> int:
    """3행부터 A열(번호)이 있는 마지막 행 반환. 데이터 없으면 2 반환."""
    last = 2
    for row_idx in range(3, ws.max_row + 1):
        if ws.cell(row_idx, 1).value is not None:
            last = row_idx
    return last


def read_last_cumulative_and_max_no_xlsx(path: Path) -> tuple[float, int]:
    """손익(N열) 값 합산으로 누적 PnL 재계산, 최대 번호 반환."""
    if not path.exists():
        return 0.0, 0
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        max_no = 0
        cum = 0.0
        pnl_col = _COL["손익"]
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] is None:
                continue
            try:
                max_no = max(max_no, int(row[0]))
            except Exception:
                continue
            pnl_raw = row[pnl_col - 1]
            if pnl_raw is not None:
                try:
                    cum += float(pnl_raw)
                except Exception:
                    pass
        return cum, max_no
    except Exception:
        return 0.0, 0


def append_result_xlsx_rows(
    path: Path,
    rows: List[Dict[str, Any]],
    symbol_names: Dict[str, str],
    kis_symbol_names: Optional[Dict[str, str]] = None,
    strategy_mode: Optional[int] = None,
) -> None:
    """result.xlsx에 rows를 이어쓰기. 파일 없으면 신규 생성."""
    if not rows:
        return
    ensure_result_xlsx(path)
    _, max_no = read_last_cumulative_and_max_no_xlsx(path)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    start_row = _find_last_data_row(ws) + 1
    kis_nm = kis_symbol_names or {}

    for i, r in enumerate(rows):
        row_num = start_row + i
        no = max_no + i + 1
        sym_raw = str(r["symbol"]).strip()
        sym = sym_raw.zfill(6) if sym_raw.isdigit() else sym_raw
        name = (symbol_names.get(sym, "") or kis_nm.get(sym, "")).strip()
        kind = str(r.get("kind", "CLOSED"))
        buy_ts = r.get("buy_ts_first") or r.get("buy_ts_last")
        buy_date = buy_ts.astimezone(KST).date() if buy_ts else None
        buy_time = buy_ts.astimezone(KST).time() if buy_ts else None

        # 종목코드: 숫자로 저장 (양식과 동일)
        sym_cell: Any = int(sym) if sym.isdigit() else sym

        # 공통 필드
        ws.cell(row_num, _COL["번호"]).value = no
        ws.cell(row_num, _COL["매수날짜"]).value = buy_date
        ws.cell(row_num, _COL["매수시간"]).value = buy_time
        ws.cell(row_num, _COL["종목코드"]).value = sym_cell
        ws.cell(row_num, _COL["종목명"]).value = name
        ws.cell(row_num, _COL["매수단가"]).value = int(round(float(r["buy_avg"])))
        ws.cell(row_num, _COL["매수수량"]).value = int(r["buy_qty"])
        ws.cell(row_num, _COL["총매수금액"]).value = int(round(float(r["buy_amt"])))
        # 누적손익: 공식으로 계산 (row 2는 헤더 → row 3부터 =R{n-1}+N{n})
        ws.cell(row_num, _COL["누적손익"]).value = f"=R{row_num - 1}+N{row_num}"

        if kind == "OPEN":
            ws.cell(row_num, _COL["매도날짜"]).value = None
            ws.cell(row_num, _COL["매도시간"]).value = None
            ws.cell(row_num, _COL["매도단가"]).value = None
            ws.cell(row_num, _COL["매도수량"]).value = None
            ws.cell(row_num, _COL["총매도금액"]).value = None
            ws.cell(row_num, _COL["손익"]).value = None
            ws.cell(row_num, _COL["수익률"]).value = None
            ws.cell(row_num, _COL["세금"]).value = int(round(float(r.get("tax", 0))))
            ws.cell(row_num, _COL["수수료"]).value = int(round(float(r.get("fee", 0))))
            ws.cell(row_num, _COL["성공률"]).value = None
            ws.cell(row_num, _COL["평균수익률"]).value = None
            ws.cell(row_num, _COL["매매전략"]).value = strategy_mode
        else:
            sell_ts = r["sell_ts"]
            pnl = float(r["pnl"])
            pnl_pct = float(r["pnl_pct"])
            ws.cell(row_num, _COL["매도날짜"]).value = sell_ts.astimezone(KST).date()
            ws.cell(row_num, _COL["매도시간"]).value = sell_ts.astimezone(KST).time()
            ws.cell(row_num, _COL["매도단가"]).value = int(round(float(r["sell_avg"])))
            ws.cell(row_num, _COL["매도수량"]).value = int(r["sell_qty"])
            ws.cell(row_num, _COL["총매도금액"]).value = int(round(float(r["sell_amt"])))
            ws.cell(row_num, _COL["손익"]).value = int(round(pnl))
            ws.cell(row_num, _COL["수익률"]).value = round(pnl_pct, 4)
            ws.cell(row_num, _COL["세금"]).value = int(round(float(r.get("tax", 0))))
            ws.cell(row_num, _COL["수수료"]).value = int(round(float(r.get("fee", 0))))
            ws.cell(row_num, _COL["성공률"]).value = f"=IF(O{row_num}>=0,1,0)"
            ws.cell(row_num, _COL["평균수익률"]).value = f"=O{row_num}/100"
            ws.cell(row_num, _COL["매매전략"]).value = strategy_mode

    wb.save(path)


def paper_trades_to_execs(paper_trades: List[Dict[str, Any]]) -> list:
    """가상 매매 기록 → Exec 리스트 변환 (fifo_sell_to_round_trips 입력용)."""
    from core.result_csv import Exec

    execs = []
    for t in paper_trades:
        qty = int(t["qty"])
        price = int(t["price"])
        sym = str(t["symbol"]).strip().zfill(6)
        execs.append(
            Exec(
                ts=t["ts"],
                side=t["side"],
                symbol=sym,
                qty=qty,
                amount=float(qty * price),
                fee=0.0,
                tax=0.0,
                avg_px=float(price),
            )
        )
    execs.sort(key=lambda e: (e.ts, 0 if e.side == "BUY" else 1))
    return execs
